"""Genera il report Excel con lo stesso formato dell'app Flutter."""
import calendar
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MESI_IT = ["", "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
           "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

K_COMUNI = ["Massa", "Sorrento", "Sant'Agnello", "Piano", "Meta", "Vico"]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, italic=False, color="FF000000", size=10) -> Font:
    return Font(bold=bold, italic=italic, color=color, size=size)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(style="thin", color="FF9E9E9E") -> Border:
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


THICK_BORDER = _border("medium", "FF424242")
THIN_BORDER = _border("thin", "FF9E9E9E")

BLU = "FF1565C0"
BLU_SCURO = "FF0D3C7A"
GIALLO = "FFFFD700"
VERDE = "FF2E7D32"
BIANCO = "FFFFFFFF"
NERO = "FF000000"
GRIGIO_CHIARO = "FFF0F0F0"
GRIGIO_RIGA = "FFD8D8D8"
GRIGIO_HEADER = "FFD0D0D0"
ARANCIO = "FFE65100"
ROSSO = "FFB71C1C"
VIOLA = "FF4A148C"


def _style(ws, cell_ref, value=None, bg=BIANCO, fg=NERO,
           bold=False, italic=False, h_align="center", wrap=True, border=THIN_BORDER):
    cell = ws[cell_ref]
    if value is not None:
        cell.value = value
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, italic=italic, color=fg)
    cell.alignment = _align(h=h_align, wrap=wrap)
    cell.border = border


def _fmt(v: float) -> str:
    if v == 0:
        return ""
    return str(int(v)) if v % 1 == 0 else f"{v:.1f}"


def generate_excel(operator, year: int, month: int, entries: list, comuni: list) -> bytes:
    wb = Workbook()

    # ── FOGLIO 1: Report ore giornaliere ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Report"

    month_name = f"{MESI_IT[month].upper()} {year}"
    days_in_month = calendar.monthrange(year, month)[1]

    # Larghezze colonne
    col_widths = [10, 22, 24, 20, 12, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Riga 1: Titolo cooperativa
    ws1.merge_cells("A1:G1")
    ws1.row_dimensions[1].height = 24
    _style(ws1, "A1", "Cooperativa Sociale Oltre i sogni a r.l. ONLUS",
           bg=BLU, fg=BIANCO, bold=True, italic=True, border=THICK_BORDER)

    # Riga 2: Nome operatore
    ws1.row_dimensions[2].height = 18
    _style(ws1, "A2", "NOME E COGNOME OPERATORE",
           bg=GRIGIO_HEADER, bold=True, h_align="left")
    ws1.merge_cells("B2:G2")
    _style(ws1, "B2", operator.full_name, bg=GRIGIO_HEADER, h_align="left")

    # Riga 3: Mese
    ws1.row_dimensions[3].height = 18
    _style(ws1, "A3", "MESE DI RIFERIMENTO", bg=GRIGIO_HEADER, bold=True, h_align="left")
    ws1.merge_cells("B3:G3")
    _style(ws1, "B3", month_name, bg=GRIGIO_HEADER, h_align="left")

    # Riga 4: Ambito
    ws1.merge_cells("A4:G4")
    ws1.row_dimensions[4].height = 18
    _style(ws1, "A4", "Servizi Ambito Na3 – Penisola",
           bg=BLU, fg=BIANCO, bold=True, italic=True, border=THICK_BORDER)

    # Riga 5: Intestazioni
    ws1.row_dimensions[5].height = 32
    headers = [
        ("A5", "GIORNO"), ("B5", "ORE SERVIZI\nMEMOFAST"),
        ("C5", "ORE\nPRIVATI"),
        ("D5", "ORE\nSOSTITUZIONI"), ("E5", "ORE\nFERIE"),
        ("F5", "ORE\nMALATTIA"), ("G5", "ORE\nLEGGE 104"),
    ]
    for ref, label in headers:
        _style(ws1, ref, label, bg=BLU_SCURO, fg=BIANCO, bold=True, border=THICK_BORDER)

    # Mappa entries per data
    entry_map = {e.date: e for e in entries}

    # Righe giorni
    tot = {k: 0.0 for k in ["ore_memofast", "ore_pulmino", "ore_sostituzioni",
                              "ore_malattia", "ore_legge104"]}
    tot_ferie_giorni = 0
    tot_ferie_ore    = 0.0

    for day in range(1, 32):
        row = day + 5
        ws1.row_dimensions[row].height = 16

        if day > days_in_month:
            row_bg = GRIGIO_RIGA
            _style(ws1, f"A{row}", day, bg=row_bg)
            for col in "BCDEFG":
                _style(ws1, f"{col}{row}", "", bg=row_bg)
            continue

        row_bg = GRIGIO_CHIARO if day % 2 == 0 else BIANCO
        _style(ws1, f"A{row}", day, bg=row_bg)

        from datetime import date as date_type
        d = date_type(year, month, day)
        e = entry_map.get(d)

        if e:
            for col, field, color in [
                ("B", "ore_memofast", NERO),
                ("C", "ore_pulmino", NERO),
                ("D", "ore_sostituzioni", NERO),
                ("F", "ore_malattia", ROSSO),
                ("G", "ore_legge104", VIOLA),
            ]:
                v = getattr(e, field)
                _style(ws1, f"{col}{row}", _fmt(v), bg=row_bg,
                       fg=color if v > 0 else NERO,
                       bold=(v > 0 and col in "EFG"))
                tot[field] += v
            # Ferie: giornata intera = "G", ore specifiche = numero
            ferie_val = e.ore_ferie
            if ferie_val == -1.0:
                _style(ws1, f"E{row}", "G", bg=row_bg, fg=ARANCIO, bold=True)
                tot_ferie_giorni += 1
            elif ferie_val > 0:
                _style(ws1, f"E{row}", _fmt(ferie_val), bg=row_bg, fg=ARANCIO, bold=True)
                tot_ferie_ore += ferie_val
            else:
                _style(ws1, f"E{row}", "", bg=row_bg)
            # Malattia: sempre giornata intera = "G"
            mal_val = e.ore_malattia
            if mal_val > 0:
                _style(ws1, f"F{row}", "G", bg=row_bg, fg=ROSSO, bold=True)
            else:
                _style(ws1, f"F{row}", "", bg=row_bg)
        else:
            for col in "BCDEFG":
                _style(ws1, f"{col}{row}", "", bg=row_bg)

    # Riga TOTALE ORE
    ws1.row_dimensions[37].height = 18
    _style(ws1, "A37", "TOTALE ORE", bg=GIALLO, bold=True, h_align="left", border=THICK_BORDER)
    for col, field, color in [
        ("B", "ore_memofast", NERO), ("C", "ore_pulmino", NERO),
        ("D", "ore_sostituzioni", NERO),
        ("F", "ore_malattia", ROSSO), ("G", "ore_legge104", VIOLA),
    ]:
        v = tot[field]
        _style(ws1, f"{col}37", _fmt(v), bg=GIALLO, fg=color, bold=True, border=THICK_BORDER)
    # Totale ferie
    if tot_ferie_giorni > 0 and tot_ferie_ore > 0:
        ferie_tot_label = f"{tot_ferie_giorni}G+{_fmt(tot_ferie_ore)}"
    elif tot_ferie_giorni > 0:
        ferie_tot_label = f"{tot_ferie_giorni} G"
    elif tot_ferie_ore > 0:
        ferie_tot_label = _fmt(tot_ferie_ore)
    else:
        ferie_tot_label = ""
    _style(ws1, "E37", ferie_tot_label, bg=GIALLO, fg=ARANCIO, bold=True, border=THICK_BORDER)
    # Totale malattia giorni
    tot_mal_giorni = sum(1 for e in entries if e.ore_malattia > 0)
    mal_tot_label = f"{tot_mal_giorni} G" if tot_mal_giorni > 0 else ""
    _style(ws1, "F37", mal_tot_label, bg=GIALLO, fg=ROSSO, bold=True, border=THICK_BORDER)

    # Riga TOTALE COMPLESSIVO
    ws1.merge_cells("A38:D38")
    ws1.row_dimensions[38].height = 18
    _style(ws1, "A38", "TOTALE COMPLESSIVO MESE",
           bg=VERDE, fg=BIANCO, bold=True, border=THICK_BORDER)
    totale = tot["ore_memofast"] + tot["ore_pulmino"] + tot["ore_sostituzioni"]
    ws1.merge_cells("E38:G38")
    _style(ws1, "E38", _fmt(totale), bg=VERDE, fg=BIANCO, bold=True, border=THICK_BORDER)

    # Legenda
    ws1.merge_cells("A39:G39")
    ws1.row_dimensions[39].height = 14
    _style(ws1, "A39",
           "LEGGENDA:   E = Ore Ferie   |   F = Ore Malattia   |   G = Ore Legge 104",
           bg=BLU, fg=BIANCO, italic=True, h_align="left")

    # ── FOGLIO 2: Servizi per Comune ─────────────────────────────────────────
    ws2 = wb.create_sheet("Servizi Comuni")

    col_widths2 = [16, 10, 10, 10, 10, 10, 14, 10, 14]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Riga 1
    ws2.merge_cells("A1:I1")
    ws2.row_dimensions[1].height = 22
    _style(ws2, "A1", "DETTAGLIO SERVIZI PER COMUNE – Na3 PENISOLA",
           bg=BLU_SCURO, fg=BIANCO, bold=True, border=THICK_BORDER)

    # Riga 2
    ws2.row_dimensions[2].height = 16
    _style(ws2, "A2", "NOME E COGNOME OPERATORE", bg=GRIGIO_HEADER, bold=True, h_align="left")
    ws2.merge_cells("B2:I2")
    _style(ws2, "B2", operator.full_name, bg=GRIGIO_HEADER, h_align="left")

    # Riga 3
    ws2.row_dimensions[3].height = 16
    _style(ws2, "A3", "MESE DI RIFERIMENTO", bg=GRIGIO_HEADER, bold=True, h_align="left")
    ws2.merge_cells("B3:I3")
    _style(ws2, "B3", month_name, bg=GRIGIO_HEADER, h_align="left")

    # Riga 4
    ws2.merge_cells("A4:I4")
    ws2.row_dimensions[4].height = 16
    _style(ws2, "A4", "Servizi Ambito Na3 – Penisola",
           bg=BLU, fg=BIANCO, bold=True, italic=True)

    # Intestazioni
    ws2.row_dimensions[5].height = 30
    headers2 = [("A5", "COMUNE"), ("B5", "ADI"), ("C5", "ADA"), ("D5", "ADH"),
                ("E5", "ADM"), ("F5", "ASIA"), ("G5", "ASIA\nIstituti\nSuperiori"),
                ("H5", "CPF"), ("I5", "TOTALE\nCOMUNE")]
    for ref, label in headers2:
        _style(ws2, ref, label, bg=BLU_SCURO, fg=BIANCO, bold=True, border=THICK_BORDER)

    # Mappa comuni
    comune_map = {cs.comune: cs for cs in comuni}
    tot2 = {k: 0.0 for k in ["adi", "ada", "adh", "adm", "asia", "asia_istituti", "cpf"]}

    for i, nome_comune in enumerate(K_COMUNI):
        row = i + 6
        ws2.row_dimensions[row].height = 16
        row_bg = GRIGIO_CHIARO if i % 2 == 0 else BIANCO
        cs = comune_map.get(nome_comune)

        _style(ws2, f"A{row}", nome_comune, bg=row_bg, h_align="left")
        for col, field in [("B", "adi"), ("C", "ada"), ("D", "adh"), ("E", "adm"),
                           ("F", "asia"), ("G", "asia_istituti"), ("H", "cpf")]:
            v = getattr(cs, field) if cs else 0
            _style(ws2, f"{col}{row}", _fmt(v), bg=row_bg)
            tot2[field] += v
        totc = sum(getattr(cs, f) for f in ["adi", "ada", "adh", "adm", "asia", "asia_istituti", "cpf"]) if cs else 0
        _style(ws2, f"I{row}", _fmt(totc), bg="FFFFF9C4")

    # Riga TOTALE SERVIZIO
    tot_row = len(K_COMUNI) + 6
    ws2.row_dimensions[tot_row].height = 18
    _style(ws2, f"A{tot_row}", "TOTALE SERVIZIO", bg=GIALLO, bold=True,
           h_align="left", border=THICK_BORDER)
    for col, field in [("B", "adi"), ("C", "ada"), ("D", "adh"), ("E", "adm"),
                       ("F", "asia"), ("G", "asia_istituti"), ("H", "cpf")]:
        _style(ws2, f"{col}{tot_row}", _fmt(tot2[field]), bg=GIALLO, bold=True, border=THICK_BORDER)
    tot_servizio = sum(tot2.values())
    _style(ws2, f"I{tot_row}", _fmt(tot_servizio), bg=GIALLO, bold=True, border=THICK_BORDER)

    # TOTALE COMPLESSIVO
    comp_row = tot_row + 1
    ws2.merge_cells(f"A{comp_row}:H{comp_row}")
    ws2.row_dimensions[comp_row].height = 18
    _style(ws2, f"A{comp_row}", "TOTALE COMPLESSIVO MESE",
           bg=VERDE, fg=BIANCO, bold=True, border=THICK_BORDER)
    _style(ws2, f"I{comp_row}", _fmt(tot_servizio),
           bg=VERDE, fg=BIANCO, bold=True, border=THICK_BORDER)

    # Footer
    foot_row = comp_row + 1
    ws2.merge_cells(f"A{foot_row}:I{foot_row}")
    ws2.row_dimensions[foot_row].height = 12
    cell = ws2[f"A{foot_row}"]
    cell.value = "sede operativa: Corso Italia, 165 – 80065 Sant'Agnello NA  |  tel: 081/16558132  |  e-mail: info@oltreisogni.org  |  P.IVA: 04018761215"
    cell.font = _font(italic=True, color="FF757575", size=8)
    cell.alignment = _align(h="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
