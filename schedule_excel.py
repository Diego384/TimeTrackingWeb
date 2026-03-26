"""Genera il file Excel della griglia oraria settimanale."""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colori ────────────────────────────────────────────────────────────────────
BLU_SCURO   = "FF0D3C7A"
BLU         = "FF1565C0"
BIANCO      = "FFFFFFFF"
NERO        = "FF000000"
ARANCIO     = "FFE65100"
GRIGIO_H    = "FFD0D0D0"
GRIGIO_RIGA = "FFF5F5F5"
GIALLO      = "FFFFD700"
VERDE       = "FF2E7D32"

GIORNI_NOMI = ["", "LUNEDÌ", "MARTEDÌ", "MERCOLEDÌ", "GIOVEDÌ", "VENERDÌ", "SABATO"]

# ── Helpers di stile ──────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, italic=False, color=NERO, size=10) -> Font:
    return Font(bold=bold, italic=italic, color=color, size=size)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _side(style="thin", color="FF9E9E9E") -> Side:
    return Side(border_style=style, color=color)


def _border(style="thin", color="FF9E9E9E") -> Border:
    s = _side(style, color)
    return Border(left=s, right=s, top=s, bottom=s)


THICK = _border("medium", "FF424242")
THIN  = _border("thin", "FF9E9E9E")


def _cell(ws, ref, value=None, bg=BIANCO, fg=NERO,
          bold=False, italic=False, h="center", wrap=False, border=THIN):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.fill   = _fill(bg)
    c.font   = _font(bold=bold, italic=italic, color=fg)
    c.alignment = _align(h=h, wrap=wrap)
    c.border = border


def _fmt_ore(v: float) -> str:
    if v == 0:
        return ""
    return str(int(v)) if v % 1 == 0 else f"{v:.2f}".rstrip("0").rstrip(".")


# ── Funzione principale ───────────────────────────────────────────────────────

def generate_schedule_excel(operator, schedule, entries_by_day: dict) -> bytes:
    """
    operator        – istanza Operator
    schedule        – istanza WeeklySchedule
    entries_by_day  – dict {day_of_week: [WeeklyScheduleEntry, ...]}
                      day_of_week: 1=Lun … 6=Sab
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Griglia Oraria"

    # Larghezze colonne: A=10, B=10, C=8, D=28, E=18, F=18
    col_widths = [10, 10, 8, 28, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Riga 1: Cooperativa ───────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 22
    _cell(ws, f"A{row}",
          "Cooperativa Sociale Oltre i sogni a r.l. ONLUS",
          bg=BLU, fg=BIANCO, bold=True, italic=True, border=THICK)
    row += 1

    # ── Riga 2: Titolo ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 20
    _cell(ws, f"A{row}", "GRIGLIA ORARIA SETTIMANALE",
          bg=BLU_SCURO, fg=BIANCO, bold=True, border=THICK)
    row += 1

    # ── Riga 3: Subtitle ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 18
    _cell(ws, f"A{row}", "Servizi Ambito Na3 – Penisola",
          bg=BLU, fg=BIANCO, italic=True, border=THICK)
    row += 1

    # ── Riga 4: Nome operatore ────────────────────────────────────────────────
    ws.row_dimensions[row].height = 16
    _cell(ws, f"A{row}", "NOME E COGNOME OPERATORE",
          bg=GRIGIO_H, bold=True, h="left", border=THIN)
    ws.merge_cells(f"B{row}:F{row}")
    _cell(ws, f"B{row}", operator.full_name,
          bg=GRIGIO_H, h="left", border=THIN)
    row += 1

    # ── Riga 5: Periodo ───────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 16
    _cell(ws, f"A{row}", "PERIODO DI RIFERIMENTO",
          bg=GRIGIO_H, bold=True, h="left", border=THIN)
    ws.merge_cells(f"B{row}:F{row}")
    _cell(ws, f"B{row}", schedule.periodo_riferimento or "",
          bg=GRIGIO_H, h="left", border=THIN)
    row += 1

    # ── Giorni 1-6 ────────────────────────────────────────────────────────────
    totale_settimana = 0.0

    for dow in range(1, 7):
        day_name = GIORNI_NOMI[dow]
        day_entries = entries_by_day.get(dow, [])

        # Intestazione giorno
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 18
        _cell(ws, f"A{row}", day_name,
              bg=BLU_SCURO, fg=BIANCO, bold=True, border=THICK)
        row += 1

        # Intestazioni colonne
        ws.row_dimensions[row].height = 28
        col_headers = ["ORA INIZIO", "ORA FINE", "ORE", "UTENTE/ASSISTITO", "SERVIZIO", "COMUNE"]
        for ci, hdr in enumerate(col_headers, 1):
            ref = f"{get_column_letter(ci)}{row}"
            _cell(ws, ref, hdr, bg=BLU, fg=BIANCO, bold=True, wrap=True, border=THIN)
        row += 1

        # Righe dati
        if day_entries:
            for idx, e in enumerate(day_entries):
                ws.row_dimensions[row].height = 16
                bg = GRIGIO_RIGA if idx % 2 == 0 else BIANCO
                _cell(ws, f"A{row}", e.ora_inizio or "", bg=bg, border=THIN)
                _cell(ws, f"B{row}", e.ora_fine or "",   bg=bg, border=THIN)
                _cell(ws, f"C{row}", _fmt_ore(e.ore),    bg=bg, border=THIN)
                _cell(ws, f"D{row}", e.utente_assistito or "", bg=bg, h="left", border=THIN)
                _cell(ws, f"E{row}", e.servizio or "",   bg=bg, border=THIN)
                _cell(ws, f"F{row}", e.comune or "",     bg=bg, border=THIN)
                row += 1
        else:
            # Riga vuota placeholder
            ws.row_dimensions[row].height = 16
            for ci in range(1, 7):
                _cell(ws, f"{get_column_letter(ci)}{row}", "", bg=BIANCO, border=THIN)
            row += 1

        # Totale giorno
        totale_giorno = sum(e.ore for e in day_entries)
        totale_settimana += totale_giorno
        ws.row_dimensions[row].height = 16
        ws.merge_cells(f"A{row}:B{row}")
        _cell(ws, f"A{row}", f"Totale {day_name.capitalize()}",
              bg=ARANCIO, fg=BIANCO, bold=True, h="right", border=THIN)
        _cell(ws, f"C{row}", _fmt_ore(totale_giorno),
              bg=ARANCIO, fg=BIANCO, bold=True, border=THIN)
        for ci in range(4, 7):
            _cell(ws, f"{get_column_letter(ci)}{row}", "", bg=ARANCIO, border=THIN)
        row += 1

    # ── Riga totale settimana ─────────────────────────────────────────────────
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:B{row}")
    _cell(ws, f"A{row}", "TOTALE ORE SETTIMANA",
          bg=GIALLO, bold=True, h="right", border=THICK)
    _cell(ws, f"C{row}", _fmt_ore(totale_settimana),
          bg=GIALLO, bold=True, border=THICK)
    for ci in range(4, 7):
        _cell(ws, f"{get_column_letter(ci)}{row}", "", bg=GIALLO, border=THICK)
    row += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    ws.row_dimensions[row].height = 12
    c = ws[f"A{row}"]
    c.value = ("sede operativa: Corso Italia, 165 – 80065 Sant'Agnello NA  |  "
               "tel: 081/16558132  |  e-mail: info@oltreisogni.org  |  P.IVA: 04018761215")
    c.font      = _font(italic=True, color="FF757575", size=8)
    c.alignment = _align(h="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
